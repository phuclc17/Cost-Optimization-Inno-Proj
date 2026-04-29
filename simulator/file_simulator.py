"""
File-based Simulator
────────────────────
Đọc events từ CSV/Excel và replay tới worker.
Dùng thay thế synthetic simulator khi muốn
test với data cố định.
"""

import os
import sys
import csv
import time
import uuid
import logging
from datetime import datetime, timezone
from typing import List, Dict
import requests

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s"
)
logger = logging.getLogger(__name__)

WORKER_URL = os.getenv("WORKER_URL", "http://localhost:8000")

# ─── READERS ─────────────────────────────────────────────

def read_csv(filepath: str) -> List[Dict]:
    """Đọc events từ file CSV."""
    events = []
    with open(filepath, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            events.append(dict(row))
    logger.info(f"Loaded {len(events)} events from {filepath}")
    return events


def read_excel(filepath: str) -> List[Dict]:
    """Đọc events từ file Excel."""
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl not installed. Run:")
        logger.error("pip3 install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    events = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        event = dict(zip(headers, row))
        # Bỏ qua row trống
        if any(v is not None for v in event.values()):
            events.append(event)
    logger.info(f"Loaded {len(events)} events from {filepath}")
    return events


def read_file(filepath: str) -> List[Dict]:
    """Auto-detect file type và đọc."""
    if filepath.endswith(".csv"):
        return read_csv(filepath)
    elif filepath.endswith((".xlsx", ".xls")):
        return read_excel(filepath)
    else:
        raise ValueError(
            f"Unsupported file: {filepath}\n"
            f"Supported: .csv, .xlsx, .xls"
        )

# ─── SENDER ──────────────────────────────────────────────

def send_event(event: Dict) -> Dict:
    """Gửi 1 event tới worker, trả về result."""
    # Đảm bảo có event_id
    if not event.get("event_id"):
        event["event_id"] = str(uuid.uuid4())

    # Đảm bảo có created_at
    if not event.get("created_at"):
        event["created_at"] = datetime.now(
            timezone.utc
        ).isoformat()

    # Đảm bảo amount là số
    try:
        event["amount"] = float(event.get("amount", 0))
    except (ValueError, TypeError):
        event["amount"] = 0.0

    try:
        start = time.perf_counter()
        resp = requests.post(
            f"{WORKER_URL}/process",
            json=event,
            timeout=5.0
        )
        latency_ms = (time.perf_counter() - start) * 1000

        if resp.status_code == 200:
            return {
                "success": True,
                "latency_ms": latency_ms,
                "event_id": event["event_id"]
            }
        else:
            return {
                "success": False,
                "error": resp.text,
                "event_id": event["event_id"]
            }
    except Exception as e:
        return {
            "success": False,
            "error": str(e),
            "event_id": event.get("event_id", "unknown")
        }

# ─── MAIN SIMULATOR ──────────────────────────────────────

class FileSimulator:
    """
    Replay events từ file với tốc độ có thể điều chỉnh.
    """

    def __init__(
        self,
        filepath: str,
        rate_rps: float = 10.0,
        repeat: int = 1
    ):
        """
        filepath : đường dẫn tới CSV/Excel
        rate_rps : số events gửi mỗi giây
        repeat   : lặp lại file bao nhiêu lần
                   (repeat=0 = lặp vô hạn)
        """
        self.events   = read_file(filepath)
        self.rate_rps = rate_rps
        self.repeat   = repeat
        self.stats    = {
            "sent": 0, "success": 0,
            "error": 0, "total_latency": 0.0
        }

    def run(self):
        """Chạy replay."""
        interval  = 1.0 / self.rate_rps
        iteration = 0

        logger.info(
            f"FileSimulator started | "
            f"events={len(self.events)} | "
            f"rate={self.rate_rps} rps | "
            f"repeat={self.repeat}"
        )

        while True:
            iteration += 1
            logger.info(f"Pass {iteration}/{self.repeat or '∞'}")

            for event in self.events:
                result = send_event(event.copy())
                self.stats["sent"] += 1

                if result["success"]:
                    self.stats["success"] += 1
                    self.stats["total_latency"] += \
                        result["latency_ms"]
                else:
                    self.stats["error"] += 1
                    logger.warning(
                        f"Failed: {result.get('error')}"
                    )

                time.sleep(interval)

            # Log sau mỗi pass
            s = self.stats["success"]
            avg = (
                self.stats["total_latency"] / s
                if s > 0 else 0
            )
            logger.info(
                f"Pass {iteration} done | "
                f"sent={self.stats['sent']} | "
                f"ok={s} | "
                f"err={self.stats['error']} | "
                f"avg_lat={avg:.1f}ms"
            )

            # Dừng nếu đủ số lần lặp
            if self.repeat > 0 and iteration >= self.repeat:
                break

        logger.info(f"FileSimulator done | stats={self.stats}")
        return self.stats


# ─── ENTRY POINT ─────────────────────────────────────────
if __name__ == "__main__":
    """
    Dùng:
    python3 file_simulator.py data/sample_events.csv 10 3
                               ^filepath              ^rate ^repeat
    """
    if len(sys.argv) < 2:
        print("Usage: python3 file_simulator.py "
              "<filepath> [rate_rps] [repeat]")
        print("Example: python3 file_simulator.py "
              "data/sample_events.csv 10 3")
        sys.exit(1)

    filepath = sys.argv[1]
    rate     = float(sys.argv[2]) if len(sys.argv) > 2 else 10.0
    repeat   = int(sys.argv[3])   if len(sys.argv) > 3 else 1

    sim = FileSimulator(filepath, rate, repeat)
    sim.run()